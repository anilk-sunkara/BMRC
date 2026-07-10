from openpyxl import load_workbook

class ExcelReader:

    @staticmethod
    def get_data():

        wb = load_workbook("testdata/BMRC_Data.xlsx")
        sheet = wb["Sheet1"]

        data = []

        for row in range(2, sheet.max_row + 1):

            from_station = sheet.cell(row, 1).value
            to_station = sheet.cell(row, 2).value

            data.append((from_station, to_station))

        return data